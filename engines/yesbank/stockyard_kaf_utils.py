from copy import copy

import pandas as pd

from openpyxl.styles import Font


def copy_row_style(
    ws,
    source_row,
    target_row
):

    for col in range(
        1,
        ws.max_column + 1
    ):

        source = ws.cell(
            source_row,
            col
        )

        target = ws.cell(
            target_row,
            col
        )

        # Use fixed data-row font instead of copying the header font
        target.font = Font(
            name="Book Antiqua",
            size=10,
            bold=False,
            italic=False,
            underline=None,
            strike=False
        )

        target.alignment = copy(
            source.alignment
        )

        target.number_format = copy(
            source.number_format
        )

        target.protection = copy(
            source.protection
        )

    ws.row_dimensions[
        target_row
    ].height = ws.row_dimensions[
        source_row
    ].height

def ensure_capacity(
    ws,
    required_rows,
    start_row
):

    existing_rows = (
        ws.max_row
        - start_row
        + 1
    )

    if required_rows <= existing_rows:

        return

    rows_needed = (
        required_rows
        - existing_rows
    )

    for _ in range(
        rows_needed
    ):

        insert_at = (
            ws.max_row
            + 1
        )

        ws.insert_rows(
            insert_at
        )

        copy_row_style(
            ws,
            insert_at - 1,
            insert_at
        )


def populate_stockyard_kaf(
    ws,
    kaf_file,
    godown_name
):

    df = pd.read_excel(
        kaf_file,
        sheet_name="Final Stock",
        keep_default_na=False
    )

    df.columns = (
        df.columns
        .astype(str)
        .str.strip()
    )

    if "Godown Name" not in df.columns:

        raise Exception(
            "'Godown Name' column not found."
        )

    df = df[
        df["Godown Name"]
        .astype(str)
        .str.strip()
        ==
        godown_name.strip()
    ]

    start_row = 2

    ensure_capacity(
        ws,
        len(df),
        start_row
    )

    for idx, row in enumerate(
        df.itertuples(index=False)
    ):

        excel_row = (
            start_row
            + idx
        )

        copy_row_style(
           ws,
           2,
           excel_row
        )

        for col_idx in range(
            1,
            21
        ):

            value = row[
                col_idx - 1
            ]

            if pd.isna(
                value
            ):

                value = ""

            ws.cell(
                excel_row,
                col_idx
            ).value = value