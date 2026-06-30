from openpyxl import load_workbook
import os

def get_schedule_info(
    base_data_file,
    agency_name
):

    wb = load_workbook(
        base_data_file,
        data_only=True
    )

    ws = wb["Schedule"]

    for row in range(
        2,
        ws.max_row + 1
    ):

        vendor = ws[
            f"H{row}"
        ].value

        if (
            vendor is None
        ):

            continue

        if (
            str(vendor).strip()
            ==
            agency_name.strip()
        ):

            return {

                "audit_period":
                    ws[
                        f"BD{row}"
                    ].value,

                "collection_file":
                    str(
                        ws[
                            f"BF{row}"
                        ].value
                    ).strip(),

                "repo_file":
                    str(
                        ws[
                            f"BG{row}"
                        ].value
                    ).strip(),

                "stockyard_file":
                    str(
                        ws[
                            f"BH{row}"
                        ].value
                    ).strip()

            }

    raise Exception(
        f"'{agency_name}' not found in Schedule sheet."
    )


def get_output_filename(
    schedule_info,
    report_type
):

    if report_type == "Collection":

        filename = schedule_info[
            "collection_file"
        ]

    elif report_type == "Repo":

        filename = schedule_info[
            "repo_file"
        ]

    elif report_type == "Stockyard":

        filename = schedule_info[
            "stockyard_file"
        ]

    else:

        raise Exception(
            f"Invalid report type '{report_type}'."
        )

    filename = str(
        filename
    ).strip()

    if not filename.lower().endswith(
        ".xlsx"
    ):

        filename += ".xlsx"

    output_folder = "output"

    os.makedirs(
        output_folder,
        exist_ok=True
    )

    return os.path.join(
        output_folder,
        filename
    )