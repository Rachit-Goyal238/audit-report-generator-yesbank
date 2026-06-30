from openpyxl.styles import Alignment


def populate_agency(
    agency_ws,
    kaf_ws,
    kaf_observation_col,
    status_col,
    errors_col,
    agency_observation_col
):

    grouped = {}

    for row in range(
        2,
        kaf_ws.max_row + 1
    ):

        sr_no = kaf_ws.cell(
            row,
            1
        ).value

        observation = kaf_ws.cell(
            row,
            kaf_observation_col
        ).value

        if (
            sr_no is None
            or
            observation in (
                None,
                ""
            )
        ):

            continue

        grouped.setdefault(
            sr_no,
            []
        ).append(
            str(observation).strip()
        )

    agency_lookup = {}

    for row in range(
        2,
        agency_ws.max_row + 1
    ):

        sr_no = agency_ws.cell(
            row,
            1
        ).value

        if sr_no is not None:

            agency_lookup[
                sr_no
            ] = row

    for sr_no, observations in grouped.items():

        if sr_no not in agency_lookup:

            continue

        row = agency_lookup[
            sr_no
        ]

        unique = []

        for obs in observations:

            if obs not in unique:

                unique.append(
                    obs
                )

        count = len(
            unique
        )

        if count > 4:

            errors = ">4"

        else:

            errors = count

        observation_text = "\n".join(
            unique
        )

        # Status Complied
        agency_ws.cell(
            row,
            status_col
        ).value = "No"

        # Number of Errors
        agency_ws.cell(
            row,
            errors_col
        ).value = errors

        # Observation
        obs_cell = agency_ws.cell(
            row,
            agency_observation_col
        )

        obs_cell.value = observation_text

        obs_cell.alignment = Alignment(
            wrap_text=True,
            vertical="top"
        )